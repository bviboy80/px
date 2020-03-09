'''
ParseFixedWidth

Parse a fixed width file and export the results as a CSV



Has header and trailer records which should be included
'''

___author__ = "Shaun Thomas"
___date__ = "October 20, 2018"
__version__ = 1.1

import sys
import struct
import csv
import os
import collections
import argparse
import openpyxl
import re


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument('-o','--org', help="Original NCOA data file")
    parser.add_argument('-m','--mm', help="Data file processed in BCC MM")
    args = parser.parse_args()
    
    outputDir = os.path.dirname(args.mm)
    orgfilename = os.path.basename(args.org)
    bccfilename = os.path.basename(args.mm)
    
    baseFilename = bccfilename.split('.')[0]
    
    recordCount = 0
    
    with open(args.org, 'rb') as o:
        with open(args.mm, 'rb') as m:

            orgheader = ["Company Number", "Account Number", "Company Name",
                "Account Registration Line 1", "Account Registration Line 2",
                "Account Registration Line 3", "Account Registration Line 4",
                "Account Registration Line 5", "Account Registration Line 6", 
                "Address Line Starting Point", "Status"]
            
            output_header = ["Status", "CoNumber", "AcctNumber", "CoName", 
                "OrigNA1", "OrigNA2", "OrigNA3", "OrigNA4", "OrigNA5", 
                "OrigNA6", "FirstAddressLine", "NewNA1", "NewNA2", "NewNA3",
                "NewNA4", "NewNA5", "NewNA6", "ReturnCode", "FootnoteCode",
                "NCOALink", "NCOAMoveDate", "NCOAMoveType"]
            
            csvInMM = csv.reader(m, delimiter=",", quoting=csv.QUOTE_ALL)
            mmheader = csvInMM.next()
            
            # Sort Mail Manger Data like the Original Data to increase speed for matching
            mmDataIn = [line for line in csvInMM]
            mmDataIn.sort()
            mmRecords = len(mmDataIn)
            
            Eligible_List = []
            Ineligible_List = []
            
            orgDataDict = createOrgDataDict(o)
            parse = makeParser()
            
            for sequenceNum, mmline in enumerate(mmDataIn, start=1):
                
                # Compare and merge on Company + Account Number values in both files
                # Parse original data line into a list
                companyAcctNum = mmline[0]
                
                orgDataRow = orgDataDict[companyAcctNum]
                orgAsciiRow = replaceNonAsciiChars(orgDataRow, sequenceNum)
                orgParsedRow = [x.strip() for x in parse(orgAsciiRow)]
                
                # Get changed address from Mail Manager data
                mmRegistrationLines = [
                    mmline[mmheader.index("FullName")], mmline[mmheader.index("Company")],
                    mmline[mmheader.index("Company2")], mmline[mmheader.index("Company3")],
                    mmline[mmheader.index("Company4")], mmline[mmheader.index("Company5")]
                    ]
                mmUpAddrLines = [
                    mmline[mmheader.index("Updated Address 1")],
                    mmline[mmheader.index("Updated Address 2")],
                    mmline[mmheader.index("Updated Address 3")],
                    mmline[mmheader.index("Updated Address 4")]
                    ]
                mmUpCity = mmline[mmheader.index("City")]
                mmUpState = mmline[mmheader.index("State")]
                mmUpZip = mmline[mmheader.index("Zip 4")]
                
                # Create NEW FULL NCOA ADDRESS
                newfulladdr = [line for line in mmRegistrationLines if line.upper() not in ["", "NULL"]]
                fixRegistrationLines(orgParsedRow, mmUpAddrLines, newfulladdr, orgheader)
                newfulladdr.append("{}, {} {}".format(mmUpCity, mmUpState, mmUpZip))
                newfulladdr = padFieldsWithBlanks(newfulladdr, 6)
                
                # Concatenate all fields and write to file
                TIN_status = orgParsedRow[-1]
                
                outputLine = [TIN_status] + orgParsedRow[:-1] + newfulladdr + \
                    [
                    mmline[mmheader.index("Return Code")],
                    mmline[mmheader.index("Footnote")],
                    mmline[mmheader.index("NCOALink Return Code")],
                    mmline[mmheader.index("Move Date")],
                    mmline[mmheader.index("Move Type")]
                    ]
                    
                if TIN_status == "":  
                    Eligible_List.append(outputLine)
                else:
                    Ineligible_List.append(outputLine)
                    
                if sequenceNum == mmRecords:
                    print "\n\nMatching Complete. {} of {} records".format(sequenceNum, mmRecords)
                    print "\nOriginal File: {}".format(orgfilename)
                    print "\nBCC File: {}".format(bccfilename)
                    
                if sequenceNum % 500 == 0:
                    print "{} of {} Records matched".format(sequenceNum, mmRecords)
                
            writeDataToCSV(outputDir, output_header, Eligible_List, Ineligible_List)
            writeDataToXLSX(outputDir, output_header, Eligible_List, Ineligible_List)


            
def createOrgDataDict(file_handle):
    """ Create a lookup to match the Mail Manager data records with 
    the original data records. Create dictionary with Company-Account 
    Number as key and the data row as the value. """
    
    orgDataDict = {line[:15]:line for line in file_handle}
    return orgDataDict


def makeParser():
    """ Create a list of slice objects to parse fields 
    in a line. Accumulate the sum of the field widths 
    to get the indices for each field. """
    
    formatStr = "5s 10s 40s 38s 38s 38s 38s 38s 38s 1s 10s"
    fieldstruct = struct.Struct(formatStr)
    parse = fieldstruct.unpack_from
    return parse

    
def replaceNonAsciiChars(line, recordcount):
    """ Convert byte text to unicode chars. Replace non-ASCII,
    the "replacement", "non-breaking space" and "Broken Bar" 
    chars. Convert chars back into bytes. """
    
    char_text = line.decode("utf-8", errors='replace').replace(u'\ufffd', " ")
    replace_nbspace = char_text.replace(u'\u00A6', " ")
    replace_bknbar = replace_nbspace.replace(u'\u00A0', " ")
    latin_line = replace_bknbar.encode("latin-1")
    ascii_char = latin_line.decode("ascii", errors='replace').replace(u'\ufffd', " ")
    ascii_line = ascii_char.encode("ascii")
    if len(ascii_line) != 295:
        print "\nCharacter error found in line {}!".format(recordcount)
        print "Line is now {}!\n".format(len(ascii_line))
        print line
        raise UnicodeError
    return ascii_line
        
   
def findAddrStartPos(orgParsedRow, orgheader):
    """ Find the index in the data row where the begins address.
    Use the first address line position and the address start value. """
    
    addrLineStart = int(orgParsedRow[orgheader.index("Address Line Starting Point")])
    registrationLinesCount = addrLineStart-1                
    registrationStartPos = orgheader.index("Account Registration Line 1")                     
    addrStartPos = registrationStartPos + registrationLinesCount
    return addrStartPos   

    
def fixRegistrationLines(orgParsedRow, mmUpAddrLines, newfulladdr, orgheader):
    """ Copy any missing registration lines from original data.
        Add new updated address lines to the new full address. """
        
    addrStartPos = findAddrStartPos(orgParsedRow, orgheader)
    
    for i in range(len(mmUpAddrLines)):
        registration_Line_Is_Missing = mmUpAddrLines[i] == "" and mmUpAddrLines[:i].count("") == i
        address_Line_Not_Empty = mmUpAddrLines[i] != ""
        
        if registration_Line_Is_Missing:
            missingRegistrationLine = orgParsedRow[addrStartPos + i]
            
            # Make sure registration line is not already in the new full address
            if missingRegistrationLine not in newfulladdr:
                newfulladdr.append(missingRegistrationLine)
        
        elif address_Line_Not_Empty:            
            mmAddrLine = mmUpAddrLines[i]
            newfulladdr.append(mmAddrLine)
        else:
            continue
            
    
def padFieldsWithBlanks(addressList, intended_length):
    """ Add blanks as needed to address lists in order to 
        create the correct position in the fixed width file. """
    
    spaces_to_add = intended_length - len(addressList)   
    return addressList + ([""] * spaces_to_add)


def writeDataToCSV(outputDir, output_header, Eligible_List, Ineligible_List):
    with open(os.path.join(outputDir, "NCOA_Records.csv"), 'wb') as n:                
        # Combine lists and write to file
        csv_writer_ncoa = csv.writer(n, delimiter=",", quoting=csv.QUOTE_ALL)
        csv_writer_ncoa.writerow(output_header)
        
        combinedData = Eligible_List + Ineligible_List
        csv_writer_ncoa.writerows(combinedData)
        
def writeDataToXLSX(outputDir, output_header, Eligible_List, Ineligible_List):
    wb = openpyxl.Workbook()
    
    ws_e = wb.create_sheet("Eligible", 0)
    ws_e.append(output_header)
    for record in Eligible_List:
        ws_e.append(record)
    
    ws_i = wb.create_sheet("Ineligible", 1)   
    ws_i.append(output_header)
    for record in Ineligible_List:
        ws_i.append(record)

    excel_file = os.path.join(outputDir, "Quarterly NCOA_36_37_38.xlsx")
    wb.save(excel_file)
        
            
if __name__ == "__main__":
    sys.exit(main())
